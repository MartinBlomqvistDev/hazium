"""Adapter for the EU Pesticides Database (Regulation (EC) No 1107/2009).

The EU PPDB (`ec.europa.eu/food/plant/pesticides/eu-pesticides-database`) is
the authoritative record of which active substances are approved for plant
protection in the EU, and the legislative acts behind each decision. It is the
source of Hazium's first *real regulatory-action* labels: which substances
lost EU approval, and when. Its backing JSON API is open and unauthenticated
(base `.../backend/api/active_substance`, endpoints `/filters`, `/search`,
`/details/{id}`); see DEV_LOG 2026-07-12 for the full access map.

This adapter reads the site's own **bulk export** rather than the API: the
"Export Active substances" button produces an XLSX of every labelable
substance (id, name, CAS, status, approval date, expiration date, legislation)
in one file, which is all the label needs. The `/details/{id}` endpoint carries
richer per-act history and member-state authorisations; ingesting that is a
documented future enrichment, not required for V1's labels.

Two regulatory events are emitted per substance where dated:

* ``APPROVAL`` — from the export's *Date of approval*, for any substance that
  was ever approved (feature: how long a substance has been on the EU market).
* ``NON_RENEWAL`` — for a substance whose current status is *Not approved* and
  that carries an *Expiration of approval* date: it was approved, then its
  approval ended and was not renewed. The expiry date is the effective
  non-renewal date and the label's positive event.

Never-approved substances (the historic non-inclusion wave) carry a status but
no dates in the export; without a date there is no temporally-placeable event,
so they are counted and skipped here (their non-inclusion year lives in the
`/details` legislative acts, the deferred enrichment). They are pre-cutoff
negatives for any recent retrodetection cutoff regardless.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from hazium.models import RegulatoryEvent, RegulatoryEventKind
from hazium.resolve.ids import safe_substance_node_id

SOURCE = "eu:ppdb"
JURISDICTION = "EU"
DEFAULT_EXPORT = Path("data/raw/ActiveSubstanceExport_12-07-2026.xlsx")

STATUS_APPROVED = "Approved"
STATUS_NOT_APPROVED = "Not approved"

#: Known CAS errors in the EU PPDB bulk export, corrected at load time so that
#: CAS-keyed resolution does not silently merge two distinct substances.
#:
#: The export lists **Maneb** (real CAS 12427-38-2) under **Mancozeb**'s CAS
#: (8018-01-7). They are distinct dithiocarbamate fungicides with distinct EC
#: numbers (Maneb 235-654-8, Mancozeb 616-995-5) and distinct EU non-renewals
#: (Maneb expired 2017-01-31, Mancozeb 2021-01-04). Without this correction,
#: ``safe_substance_node_id`` maps both rows to ``substance:cas:8018-01-7`` and
#: Maneb's 2017 non-renewal is attributed to Mancozeb; since HEWB anchors
#: lead-time to the *earliest* non-renewal, Mancozeb's benchmark action date
#: becomes 2017 instead of its real 2021 date. Verified against the raw export
#: rows 2026-07-19 (DEV_LOG). Keyed on (name, exported CAS) so only the
#: mis-CAS'd Maneb row is touched, never the genuine Mancozeb row.
_CAS_CORRECTIONS: dict[tuple[str, str], str] = {
    ("Maneb", "8018-01-7"): "12427-38-2",
}

_COLUMNS = {
    "as_id": "Active Substance ID",
    "name": "Substance",
    "cas": "CAS Number",
    "approval": "Date of approval",
    "expiry": "Expiration of approval",
    "legislation": "Legislation",
}


@dataclass(frozen=True)
class _ASRow:
    """One active substance from the bulk export."""

    as_id: str
    name: str
    cas_number: str | None
    status: str
    approval_date: date | None
    expiry_date: date | None


def load_export(xlsx_path: Path) -> list[_ASRow]:
    """Parse the active-substance bulk export.

    The sheet opens with a title banner, so the header is located by scanning
    for the row containing ``Active Substance ID`` rather than assuming a fixed
    offset (banner wording changes with the export date).
    """
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header_i = next(
        i for i, r in enumerate(rows) if _COLUMNS["as_id"] in [str(c) if c else c for c in r]
    )
    header = rows[header_i]
    idx = {key: header.index(col) for key, col in _COLUMNS.items()}
    status_i = next(i for i, c in enumerate(header) if c and str(c).startswith("Status"))

    out = []
    for row in rows[header_i + 1 :]:
        as_id = row[idx["as_id"]]
        if not as_id:
            continue
        name = str(row[idx["name"]])
        cas_number = _clean_cas(row[idx["cas"]])
        if cas_number is not None:
            cas_number = _CAS_CORRECTIONS.get((name, cas_number), cas_number)
        out.append(
            _ASRow(
                as_id=str(as_id),
                name=name,
                cas_number=cas_number,
                status=str(row[status_i]) if row[status_i] else "",
                approval_date=_parse_date(row[idx["approval"]]),
                expiry_date=_parse_date(row[idx["expiry"]]),
            )
        )
    return out


def _clean_cas(value: Any) -> str | None:
    """Non-CAS entries read ``'No CAS allocated'``; treat those as absent."""
    text = str(value).strip() if value else None
    if not text or "No CAS" in text:
        return None
    return text


def _parse_date(value: Any) -> date | None:
    """Export dates are ``DD/MM/YYYY`` strings (or datetimes from openpyxl).

    Some cells hold a legislation reference instead of a date (pending
    substances); those fail to parse and yield ``None``, which correctly
    suppresses an event rather than inventing one.
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def regulatory_events_from(rows: list[_ASRow]) -> list[RegulatoryEvent]:
    """Dated EU regulatory events, resolved to canonical substance ids.

    ``known_at`` equals the event date: an approval/non-renewal is publicly
    knowable by the date it takes effect. For non-renewal this is conservative
    (the non-renewing regulation is usually published before the approval
    actually expires), which is the safe direction for temporal integrity: it
    never claims the event was knowable earlier than it provably was.
    """
    events = []
    for row in rows:
        substance_id = safe_substance_node_id(cas_number=row.cas_number, name=row.name)
        if row.approval_date:
            events.append(_event(substance_id, RegulatoryEventKind.APPROVAL, row.approval_date))
        if row.status == STATUS_NOT_APPROVED and row.expiry_date:
            events.append(_event(substance_id, RegulatoryEventKind.NON_RENEWAL, row.expiry_date))
    return events


def _event(substance_id: str, kind: RegulatoryEventKind, event_date: date) -> RegulatoryEvent:
    return RegulatoryEvent(
        substance_id=substance_id,
        kind=kind,
        jurisdiction=JURISDICTION,
        event_date=event_date,
        source=SOURCE,
        known_at=event_date,
    )
