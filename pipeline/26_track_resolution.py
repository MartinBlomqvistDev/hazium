"""Record what the register says about watchlist substances, and score the misses.

The watchlist is a set of dated, falsifiable claims, so it needs a scoreboard.
This reads the EU Pesticides Database export, stores each watchlist substance's
approval status and expiry, and compares that against the first reading ever
taken. Hits, confirmed false positives and still-open cases are reported apart,
because collapsing the third into the second is what makes an early-warning
model look worse than it is.

**The baseline is captured forward, not reconstructed.** There is exactly one
PPDB export in this repo and the snapshot archive holds no PPDB history, so
there is nothing to diff against yet. The first run writes the baseline; every
later run compares against it. Until a substance's approval expires, its row is
honestly blank rather than counted as wrong.

**A passed expiry is not a verdict.** The Commission frequently extends an
approval by a short procedural step while an assessment finishes. That looks
like survival and would score as a false positive if only status were read, so
the length of the new term decides, and an unmoved but elapsed expiry is
reported as overdue rather than resolved.

Usage:
    python pipeline/26_track_resolution.py
    python pipeline/26_track_resolution.py --variant early_warning --top 100
"""

from __future__ import annotations

import argparse
import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from hazium.benchmark.resolution import (
    ApprovalState,
    Resolution,
    TrackedEntry,
    classify,
    confirmed_precision,
    summarise,
)

ROOT = Path(__file__).parent.parent
PROCESSED = ROOT / "data" / "processed"
RAW = ROOT / "data" / "raw"

#: Column offsets in the Active Substances export, which opens with a title
#: banner so the header sits on row 3 and data begins on row 4.
COL_NAME = 1
COL_CAS = 2
COL_STATUS = 3
COL_EXPIRY = 5
FIRST_DATA_ROW = 4


def _as_date(value: object) -> date | None:
    """Coerce a register cell to a date, tolerating both native and dd/mm/yyyy."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    try:
        return datetime.strptime(str(value)[:10], "%d/%m/%Y").date()
    except ValueError:
        return None


def read_register(path: Path, observed_at: date) -> dict[str, ApprovalState]:
    """Read approval status and expiry per substance from the PPDB export."""
    workbook = load_workbook(path, read_only=True)
    states: dict[str, ApprovalState] = {}
    for row in workbook.active.iter_rows(min_row=FIRST_DATA_ROW, values_only=True):
        if not row or not row[0]:
            continue
        raw_cas = row[COL_CAS]
        if not raw_cas or "No CAS" in str(raw_cas):
            continue
        substance_id = f"substance:cas:{str(raw_cas).strip()}"
        states[substance_id] = ApprovalState(
            substance_id=substance_id,
            name=str(row[COL_NAME] or "").strip(),
            status=str(row[COL_STATUS] or "").strip(),
            expiry=_as_date(row[COL_EXPIRY]),
            observed_at=observed_at,
        )
    return states


def load_baseline(path: Path) -> dict[str, ApprovalState]:
    """Previously recorded first readings, keyed by substance id."""
    if not path.exists():
        return {}
    with path.open(encoding="utf-8") as f:
        states = [ApprovalState.model_validate_json(line) for line in f if line.strip()]
    return {s.substance_id: s for s in states}


def write_baseline(path: Path, states: dict[str, ApprovalState]) -> None:
    with path.open("w", encoding="utf-8") as f:
        for substance_id in sorted(states):
            f.write(states[substance_id].model_dump_json() + "\n")


def load_watchlist(variant: str, top: int) -> list[tuple[int, str, str]]:
    """Top-N watchlist rows as ``(rank, substance_id, name)``."""
    path = PROCESSED / f"current_watchlist_{variant}.csv"
    if not path.exists():
        raise SystemExit(f"missing {path}; run pipeline/13_current_watchlist.py first")
    rows: list[tuple[int, str, str]] = []
    with path.open(encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            rank = int(row["rank"])
            if rank > top:
                break
            rows.append((rank, row["substance_id"], row["name"]))
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--variant", default="headline", choices=("headline", "early_warning"))
    parser.add_argument("--top", type=int, default=100)
    parser.add_argument(
        "--export",
        type=Path,
        default=RAW / "ActiveSubstanceExport_12-07-2026.xlsx",
        help="EU Pesticides Database Active Substances export to read",
    )
    args = parser.parse_args()

    today = date.today()
    current = read_register(args.export, today)
    baseline_path = PROCESSED / f"resolution_baseline_{args.variant}.jsonl"
    baseline = load_baseline(baseline_path)
    watchlist = load_watchlist(args.variant, args.top)

    print(f"watchlist top {args.top}: {len(watchlist)} substances")
    print(f"register rows read: {len(current)}")
    print(f"baseline entries on file: {len(baseline)}")

    # Record a first reading for anything not yet tracked. Existing baselines
    # are never overwritten: the whole point is to preserve what was true when
    # the prediction was made.
    added = 0
    for _rank, substance_id, _name in watchlist:
        if substance_id not in baseline and substance_id in current:
            baseline[substance_id] = current[substance_id]
            added += 1
    if added:
        write_baseline(baseline_path, baseline)
        print(f"recorded {added} new baseline readings -> {baseline_path.name}")

    entries: list[TrackedEntry] = []
    missing = 0
    for rank, substance_id, name in watchlist:
        base, now = baseline.get(substance_id), current.get(substance_id)
        if base is None or now is None:
            missing += 1
            continue
        entries.append(
            TrackedEntry(substance_id=substance_id, name=name, rank=rank, baseline=base, latest=now)
        )

    counts = summarise(entries, today)
    hits, settled, precision = confirmed_precision(entries, today)

    print(f"\ntracked {len(entries)} of {len(watchlist)} ({missing} not in the register export)")
    print("\noutcome distribution:")
    for outcome in Resolution:
        if counts[outcome]:
            print(f"  {outcome.value:<14} {counts[outcome]:>4}")
    if settled:
        print(f"\nconfirmed precision on settled entries: {hits}/{settled} = {precision:.0%}")
    else:
        print("\nnothing has settled yet. No precision is reported, because 0/0 is not zero.")

    out = PROCESSED / f"resolution_{args.variant}.csv"
    with out.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "rank",
                "substance",
                "baseline_status",
                "baseline_expiry",
                "current_status",
                "current_expiry",
                "outcome",
            ]
        )
        for entry in sorted(entries, key=lambda e: e.rank):
            writer.writerow(
                [
                    entry.rank,
                    entry.name,
                    entry.baseline.status,
                    entry.baseline.expiry or "",
                    entry.latest.status,
                    entry.latest.expiry or "",
                    classify(entry, today).value,
                ]
            )
    print(f"wrote {out}")

    due = sorted(
        (
            e
            for e in entries
            if e.baseline.expiry and e.baseline.expiry <= date(today.year + 1, 12, 31)
        ),
        key=lambda e: e.baseline.expiry,
    )
    print(f"\n{len(due)} of {len(entries)} reach their expiry by end of {today.year + 1}:")
    for entry in due[:12]:
        print(f"  {entry.baseline.expiry}  rank {entry.rank:>3}  {entry.name[:46]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
