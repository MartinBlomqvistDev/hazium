"""Export the forward watchlist, its crop exposure and its resolution calendar.

The site is a static build with no access to ``data/processed/``, which is
gitignored, so this is the bridge: the CSVs written by `pipeline/13`, `25` and
`26` become one small committed JSON under ``web/data/``.

Everything else on the site is retrospective. This is the one surface making a
claim about a future that has not happened, so two things are carried alongside
the numbers rather than left to the copy:

**Each substance's approval expiry.** That date is when the Commission is
forced to decide, which is what turns a ranking into a falsifiable claim with a
deadline instead of an open-ended assertion.

**No product or brand names.** The crop is the finest granularity published. At
an average precision around 0.25 most of these substances will not be actioned,
and attaching named commercial products to a mostly-wrong list would be both
unfair and legally exposed. Counts by crop carry the same information without
naming anyone.

Usage:
    python pipeline/27_export_watchlist_site_data.py
"""

from __future__ import annotations

import csv
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from hazium.models import SalesRecord, Substance
from hazium.resolve.names import SubstanceResolver, resolve_sales_records

ROOT = Path(__file__).parent.parent
PROCESSED = ROOT / "data" / "processed"
RAW = ROOT / "data" / "raw"
SITE_DATA = ROOT / "web" / "data" / "watchlist.json"

#: The EU active-substances export, used to keep the sales denominator to plant
#: protection actives.
PPP_EXPORT = RAW / "ActiveSubstanceExport_12-07-2026.xlsx"

VARIANT = "headline"

#: How many watchlist ranks to publish. Precision holds a plateau to about k=50
#: on cutoffs old enough to have resolved (0.68-0.78) and is still 0.57-0.62 at
#: k=100, while recall roughly doubles, so the exposure map uses the wider band.
TOP = 100

#: Crops shown on the site. The register's catch-all labels ("fruit
#: (unspecified)", "berries (unspecified)") are real but say little, so they are
#: dropped here rather than padding a table a reader is meant to scan.
CROP_EXCLUDE = frozenset({"fruit (unspecified)", "berries (unspecified)", "cereals (unspecified)"})

#: Crops needing at least this many approved products before they are shown. A
#: crop with three products cannot support a percentage anyone should read.
MIN_PRODUCTS = 10


def _plant_protection_ids() -> set[str]:
    """CAS ids of substances the EU register lists as plant protection actives."""
    workbook = load_workbook(PPP_EXPORT, read_only=True)
    ids: set[str] = set()
    for row in workbook.active.iter_rows(min_row=4, values_only=True):
        if row and row[0] and row[2] and "No CAS" not in str(row[2]):
            ids.add(f"substance:cas:{str(row[2]).strip()}")
    return ids


def swedish_sales() -> tuple[dict[str, float], dict[str, int], int, int]:
    """Latest-year Swedish tonnage per substance, and its rank among peers.

    Raw tonnage says nothing on its own: the median plant protection active
    sells 0.2 tonnes a year in Sweden while the largest sells 783, so a number
    without a rank beside it is unreadable.

    The denominator is restricted to plant protection actives on purpose. KemI's
    sales file covers biocides too, and they dominate it: creosote alone is
    2,761 tonnes, against 1,992 tonnes for every plant protection active
    combined. Ranking a fungicide inside that total would be comparing it
    against wood preservative.

    Returns:
        ``(tonnes_by_id, rank_by_id, n_ranked, year)``.
    """

    def _load(path: Path, model):
        with path.open(encoding="utf-8") as f:
            return [model.model_validate_json(line) for line in f]

    # The sales file keys substances by Swedish name, so it has to go through
    # the same resolver the model uses before it can join to anything.
    sales = resolve_sales_records(
        _load(PROCESSED / "kemi_sales.jsonl", SalesRecord),
        SubstanceResolver(_load(PROCESSED / "kemi_register_substances.jsonl", Substance)),
    )
    ppp = _plant_protection_ids()
    year = max(s.year for s in sales)
    tonnes: dict[str, float] = {}
    for record in sales:
        if record.year != year or record.tonnes_active_substance is None:
            continue
        if record.substance_id not in ppp:
            continue
        tonnes[record.substance_id] = (
            tonnes.get(record.substance_id, 0.0) + record.tonnes_active_substance
        )
    ordered = sorted(tonnes.items(), key=lambda kv: -kv[1])
    ranks = {sid: i + 1 for i, (sid, _t) in enumerate(ordered)}
    return tonnes, ranks, len(ordered), year


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        raise SystemExit(f"missing {path}; run the pipeline that writes it first")
    with path.open(encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def main() -> int:
    watchlist = _read_csv(PROCESSED / f"current_watchlist_{VARIANT}.csv")[:TOP]
    crops = _read_csv(PROCESSED / f"crop_exposure_{VARIANT}.csv")
    by_substance = _read_csv(PROCESSED / f"crop_exposure_{VARIANT}_by_substance.csv")
    resolution = _read_csv(PROCESSED / f"resolution_{VARIANT}.csv")

    tonnes, sales_rank, n_ranked, sales_year = swedish_sales()
    crops_of = {r["substance"]: r["crops"] for r in by_substance}
    expiry_of = {r["substance"]: r["baseline_expiry"] for r in resolution}
    outcome_of = {r["substance"]: r["outcome"] for r in resolution}

    entries = []
    for row in watchlist:
        name = row["name"]
        crop_list = [c.strip() for c in (crops_of.get(name) or "").split(";") if c.strip()]
        entries.append(
            {
                "rank": int(row["rank"]),
                "name": name,
                "in_sweden": row["in_kemi_sweden_register"] == "True",
                "crops": crop_list,
                "expiry": expiry_of.get(name) or None,
                "outcome": outcome_of.get(name) or "untracked",
                "tonnes": round(tonnes[row["substance_id"]], 1)
                if row["substance_id"] in tonnes
                else None,
                "sales_rank": sales_rank.get(row["substance_id"]),
            }
        )

    crop_rows = [
        {
            "crop": r["crop"],
            "products": int(r["approved_products"]),
            "flagged": int(r["products_with_watchlist_substance"]),
            "percent": int(r["percent"]),
        }
        for r in crops
        if r["crop"] not in CROP_EXCLUDE and int(r["approved_products"]) >= MIN_PRODUCTS
    ]

    # Resolution calendar: how many entries reach a forced decision, by year.
    dated = sorted(e["expiry"] for e in entries if e["expiry"])
    calendar: list[dict[str, int]] = []
    running = 0
    for year in sorted({int(d[:4]) for d in dated}):
        count = sum(1 for d in dated if int(d[:4]) == year)
        running += count
        calendar.append({"year": year, "count": count, "cumulative": running})

    # The base rate must come from distinct products, not from summing the
    # per-crop table: a product approved for wheat, barley and rye sits in three
    # rows, so the columns cannot be added. Here the wrong method landed close
    # (35% against 37%) because the double-counting is spread fairly evenly, but
    # that is luck rather than a reason to keep it: the error scales with how
    # much crop overlap the register happens to carry.
    summary = _read_csv(PROCESSED / f"crop_exposure_{VARIANT}_summary.csv")[0]
    with_crop = int(summary["products_with_a_crop"])
    flagged = int(summary["products_with_watchlist_substance"])

    payload = {
        "variant": VARIANT,
        "generated": date.today().isoformat(),
        "top": TOP,
        "tracked": len(dated),
        "on_market": sum(1 for e in entries if e["crops"]),
        "calendar": calendar,
        "crops": crop_rows,
        # Keep anything with a crop, a deadline or a sales volume. Dropping the
        # volume-only rows would have hidden acetic acid, the third largest
        # plant protection seller in Sweden and rank 62 here, which is exactly
        # the kind of entry a reader should see: a basic substance the model
        # rates highly, carrying no expiry because its approval is open-ended.
        "entries": [e for e in entries if e["crops"] or e["expiry"] or e["tonnes"]],
        "with_sales": sum(1 for e in entries if e["tonnes"] is not None),
        "base_rate_percent": round(flagged / with_crop * 100) if with_crop else 0,
        "sales_year": sales_year,
        "sales_ranked": n_ranked,
    }

    SITE_DATA.parent.mkdir(parents=True, exist_ok=True)
    SITE_DATA.write_text(json.dumps(payload, indent=1, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"watchlist entries published: {len(payload['entries'])} of top {TOP}")
    print(f"with a Swedish crop use:     {payload['on_market']}")
    print(f"with a dated expiry:         {payload['tracked']}")
    print(f"crops shown:                 {len(crop_rows)}")
    print("\nresolution calendar:")
    for row in calendar[:6]:
        print(f"  {row['year']}  {row['count']:>3} decided   cumulative {row['cumulative']:>3}")
    print(f"\nwrote {SITE_DATA.relative_to(ROOT)} ({SITE_DATA.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
